import openpyxl
from datetime import datetime

from .database import SessionLocal
from .models import Institution


EXCEL_FILE = "data/institutions/College-Affiliated College.xlsx"


def clean(value):
    if value is None:
        return None

    value = str(value).strip()

    if not value:
        return None

    return value


def clean_year(value):
    value = clean(value)

    if not value:
        return None

    try:
        return int(float(value))
    except (ValueError, TypeError):
        return None


def import_aishe():
    print("Opening AISHE Excel file...")

    workbook = openpyxl.load_workbook(
        EXCEL_FILE,
        read_only=True,
        data_only=True,
    )

    sheet = workbook[workbook.sheetnames[0]]

    rows = sheet.iter_rows(
        min_row=4,
        values_only=True,
    )

    db = SessionLocal()

    inserted = 0
    updated = 0
    skipped = 0

    try:
        for row_number, row in enumerate(rows, start=4):

            if len(row) < 12:
                skipped += 1
                continue

            (
                aishe_code,
                name,
                state,
                district,
                website,
                year_of_establishment,
                location,
                college_type,
                management,
                university_aishe_code,
                university_name,
                university_type,
            ) = row[:12]

            aishe_code = clean(aishe_code)
            name = clean(name)
            state = clean(state)

            if not aishe_code or not name or not state:
                skipped += 1
                continue

            existing = (
                db.query(Institution)
                .filter(
                    Institution.aishe_code == aishe_code
                )
                .first()
            )

            data = {
                "aishe_code": aishe_code,
                "name": name,
                "state": state,
                "district": clean(district),
                "website": clean(website),
                "established_year": clean_year(
                    year_of_establishment
                ),
                "location": clean(location),
                "institution_type": clean(college_type),
                "management": clean(management),
                "university_aishe_code": clean(
                    university_aishe_code
                ),
                "university": clean(university_name),
                "university_type": clean(university_type),
                "source": "AISHE",
            }

            if existing:
                for key, value in data.items():
                    setattr(existing, key, value)

                updated += 1

            else:
                institution = Institution(
                    **data,
                    created_at=datetime.utcnow(),
                )

                db.add(institution)
                inserted += 1

            if (inserted + updated) % 1000 == 0:
                db.commit()
                print(
                    f"Processed {inserted + updated} records..."
                )

        db.commit()

        print()
        print("AISHE import completed.")
        print(f"Inserted: {inserted}")
        print(f"Updated:  {updated}")
        print(f"Skipped:  {skipped}")

    except Exception:
        db.rollback()
        raise

    finally:
        db.close()
        workbook.close()


if __name__ == "__main__":
    import_aishe()